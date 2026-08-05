"""Genera los XLSX de la documentación v4.

- USUARIOS_ROLES_Y_CODIGOS.xlsx — requiere docs/_users_export.json
  (generarlo antes con: node scripts/export_users_json.js). Contiene
  credenciales reales: NO versionar el resultado.
- MODULOS_Y_PORTALES.xlsx — catálogo estático de módulos y portales.

Uso:  python build_xlsx_v4.py
"""

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent

NAVY = "0E2A47"
PRIMARY = "1F4E8C"
ACCENT = "E3A808"
LIGHT = "F4F6FA"
WARN = "FDE2E2"

HEADER_FILL = PatternFill("solid", fgColor=PRIMARY)
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
ALT_FILL = PatternFill("solid", fgColor="F7F9FC")
WARN_FILL = PatternFill("solid", fgColor=WARN)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
NOTE_FONT = Font(color="5E6B7A", italic=True, size=10)
THIN = Side(style="thin", color="DDE3EA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def add_sheet(wb, name, title, headers, rows, widths=None, note=None, highlight=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=title)
    c.fill = TITLE_FILL
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    for j in range(2, len(headers) + 1):
        ws.cell(row=1, column=j).fill = TITLE_FILL

    r0 = 3
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=r0, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[r0].height = 20

    for i, row in enumerate(rows):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r0 + 1 + i, column=j, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if i % 2 == 1:
                c.fill = ALT_FILL
            if highlight and highlight(row):
                c.fill = WARN_FILL

    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w

    if note:
        nr = r0 + len(rows) + 2
        ws.cell(row=nr, column=1, value=note).font = NOTE_FONT

    ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
    return ws


# ───────────────────────────────────────────────────────────────────────────
# Catálogo de módulos y roles (mantener sincronizado con
# frontend/components/SideNav.tsx y frontend/app/(main)/admin/usuarios/page.tsx)
# ───────────────────────────────────────────────────────────────────────────

MODULOS = [
    ("Dashboard", "Dashboard Comercial", "/dashboard/comercial", "dashboard.comercial", "Indicadores comerciales del evento"),
    ("Dashboard", "Dashboard Operacional", "/dashboard/operacional", "dashboard.operacional", "Indicadores operativos del día"),
    ("Registro", "Registro Evento", "/registro/eventos", "registro.eventos", "Creación y configuración de eventos"),
    ("Registro", "Inscripción Participantes", "/registro/participantes", "registro.participantes", "Alta y validación de participantes (atletas, VIP, jefes)"),
    ("Registro", "Proveedores", "/registro/proveedores", "registro.participantes", "Proveedores y sus participantes (choferes, staff) con documentos"),
    ("Registro", "Clientes", "/clientes", "clientes", "Tipos de cliente y su configuración"),
    ("Operación · Arribos & Salidas", "AND", "/operacion/and", "operacion.and", "Planilla de llegadas y salidas (Arrivals & Departures)"),
    ("Operación · Arribos & Salidas", "Cumplimiento AND", "/operacion/cumplimiento-and", "operacion.cumplimiento", "Control del cumplimiento del plan AND"),
    ("Operación · Arribos & Salidas", "Monitor de Vuelos", "/operations/flights", "operacion.and", "Vuelos en tiempo real, transfers vinculados y timeline clickeable"),
    ("Operación · Arribos & Salidas", "Monitoreo de Salidas", "/operacion/salidas", "operacion.and", "Seguimiento de salidas de delegaciones"),
    ("Operación · Transporte", "Operatividad Diaria", "/operations/daily-transport", "operacion.viajes", "Gestión del día: cobertura, asignación automática, carga XLSX"),
    ("Operación · Transporte", "Calendario Operacional", "/sports-calendar", "calendario", "Calendario deportivo/operacional sincronizado"),
    ("Operación · Transporte", "Tracking de Viajes", "/operations/vehicle-positions", "operacion.tracking", "Mapa en tiempo real de vehículos en viaje"),
    ("Operación · Transporte", "Viajes", "/operations/trips", "operacion.viajes", "Listado y gestión completa de viajes"),
    ("Operación · Transporte", "Solicitudes (T1/VIP)", "/operations/trip-requests", "operacion.viajes", "Solicitudes de vehículo de VIPs y T1"),
    ("Operación · Transporte", "Flota (disponibilidad)", "/operations/fleet", "operacion.viajes", "Disponibilidad de vehículos y flota VIP/T1"),
    ("Operación · Transporte", "Panel Conductores", "/operations/driver-heatmap", "operacion.viajes", "Heatmap de carga horaria por conductor"),
    ("Operación · Transporte", "Monitoreo de Conductores", "/operations/driver-monitoring", "operacion.tracking", "Presencia y última posición de cada conductor"),
    ("Operación · Transporte", "Monitoreo VIP", "/operations/vip-monitoring", "operacion.tracking", "Tracking permanente de VIPs con ficha y timeline"),
    ("Operación · Transporte", "Panel Financiero", "/operations/transport-finance", "operacion.viajes", "Ingreso, costo y margen de la operación de transporte"),
    ("Operación · Sede", "Sede", "/sede", "sede", "Sedes (venues) del evento"),
    ("Operación · Hotelería", "Tracking Hotelería", "/operations/hotel-tracking", "hoteleria.tracking", "Estado de check-in/out y ocupación"),
    ("Operación · Hotelería", "Hoteles / Villa Panamericana", "/masters/accommodations", "hoteleria.hoteles", "Maestro de hoteles"),
    ("Operación · Hotelería", "Habitaciones", "/masters/hotel-rooms", "hoteleria.habitaciones", "Inventario de habitaciones"),
    ("Operación · Hotelería", "Asignaciones Hotel", "/operations/hotel-assignments", "hoteleria.asignaciones", "Asignación de habitaciones a participantes"),
    ("Operación · Hotelería", "Gestión de llaves", "/operations/hotel-keys", "hoteleria.llaves", "Entrega y devolución de llaves"),
    ("Operación · Hotelería", "Reserva de salones", "/operations/salones", "hoteleria.llaves", "Reserva de salones y espacios comunes"),
    ("Operación · Hotelería", "Reserva de Extras", "/operations/hotel-extras", "hoteleria.llaves", "Extras de hotel (coffee break, equipamiento)"),
    ("Operación · Alimentación", "Tipos de Alimentación", "/operations/food/tipos", "alimentacion.general", "Tipos y restricciones dietarias"),
    ("Operación · Alimentación", "Desayuno", "/operations/food/desayuno", "alimentacion.general", "Menús de desayuno"),
    ("Operación · Alimentación", "Almuerzos", "/operations/food/almuerzos", "alimentacion.general", "Menús de almuerzo"),
    ("Operación · Alimentación", "Cenas", "/operations/food/cenas", "alimentacion.general", "Menús de cena"),
    ("Operación · Alimentación", "Lugares de comida", "/operations/food/lugares", "alimentacion.general", "Ubicaciones de alimentación"),
    ("Operación · Salud", "Salud", "/health", "salud", "Fichas de salud de participantes con detalle individual"),
    ("Operación · Asistencia", "Centro de Incidencias", "/operations/support-chats", "(siempre visible)", "Chats de asistencia de todos los portales"),
    ("Operación · Workforce", "Staff & Voluntarios", "/operations/workforce", "operacion.viajes", "Personal operativo, productos y entregas"),
    ("Deportes", "Planificación deportiva", "/deportes", "deportes", "Deportes, disciplinas, pruebas y cupos"),
    ("Deportes", "Premiaciones", "/deportes/premiaciones", "deportes", "Ceremonias de premiación y equipos VIP"),
    ("Beneficios", "Administrar beneficios", "/operations/coupons", "operacion.viajes", "Cupones, partners (comercios) y reportería de canjes"),
    ("Beneficios", "Portal Partner", "/portal/partner", "portales", "Acceso directo al portal del comercio"),
    ("Acreditación", "Acreditación", "/accreditations", "acreditaciones", "Flujo de acreditación y credenciales (incluye conductores de proveedor)"),
    ("Portales", "Portal de usuario", "/portal/user", "portales", "Acceso directo al portal del participante"),
    ("Portales", "Portal Conductor", "/portal/conductor", "portales", "Acceso directo al portal del conductor"),
    ("Portales", "Solicitud de vehículo", "/portal/vehicle-request", "portales", "Acceso directo al portal VIP"),
    ("Portales", "Control de Acceso", "/portal/access-control", "portales", "Acceso directo al escáner de credenciales"),
    ("Administración", "Gestión de Usuarios", "/admin/usuarios", "admin.usuarios", "Usuarios admin, roles y módulos"),
    ("Administración", "Notificaciones Push", "/admin/notificaciones", "admin.notificaciones", "Envío manual de push a dispositivos registrados"),
    ("Administración", "Acciones de SofIA", "/operations/sofia-actions", "(siempre visible)", "Auditoría de operaciones del asistente de IA"),
    ("Otros", "Inicio guiado", "/inicio-guiado", "(siempre visible)", "Recorrido de bienvenida por rol"),
    ("Otros", "Ayuda", "/ayuda", "(siempre visible)", "Centro de ayuda y Cuaderno de Cargo"),
]

PORTALES = [
    ("Portal de usuario", "/portal/user", "Participantes TA (atletas y jefes de delegación)", "Código de 6 caracteres",
     "Itinerario, actividades/viajes, calendario deportivo, sedes, alimentación, premiaciones, cupones, ficha de salud, credencial QR, cuenta (con eliminar cuenta)"),
    ("Solicitud de Vehículo (VIP)", "/portal/vehicle-request", "Participantes tipo VIP (redirigido automáticamente desde /portal/user)", "Código de 6 caracteres",
     "Solicitar vehículo, seguimiento con mapa, premiaciones, cupones, sedes/hoteles/comida, SofIA, credencial, cuenta (con eliminar cuenta)"),
    ("Portal Conductor", "/portal/conductor", "Conductores propios y de proveedor (isDriver)", "Código de 6 caracteres",
     "Bandeja de viajes, flujo del viaje (iniciar → recoger con código → finalizar), tracking GPS, reportes/fotos, documentos, credencial, cuenta (con eliminar cuenta)"),
    ("Control de Acceso", "/portal/access-control", "Staff de proveedores tipo Staff", "Código de 6 caracteres",
     "Escáner QR de credenciales (verde/rojo/amarillo) con bitácora de accesos; eliminar cuenta"),
    ("Portal Partner", "/portal/partner", "Comercios aliados de beneficios", "Código de comercio + PIN (sesión de 12 h)",
     "Escanear/canjear cupones, código manual, estadísticas de canje; eliminar cuenta"),
    ("Login móvil", "/m/login", "Participantes y conductores desde la app nativa", "Código de 6 caracteres",
     "Resuelve automáticamente el portal según el código; recuperación por correo en /m/recover"),
    ("Credencial sin login", "/credencial", "Operación en terreno", "Enlace directo",
     "Descarga del PDF de la credencial completa sin iniciar sesión"),
]

ROLES = [
    ("Administrador", "Todos los módulos (incluida Administración)"),
    ("Supervisor", "Todo excepto el grupo Administración"),
    ("Coordinador", "Todo excepto Dashboard y Administración"),
    ("Operador", "Operación, Transporte, Hotelería, Alimentación y Acreditaciones"),
    ("Visualizador", "Solo Dashboard y Registro"),
]


def build_modulos():
    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(
        wb, "Módulos Admin", "Seven Arena — Módulos de administración",
        ["Grupo", "Módulo", "Ruta", "ID de permiso", "Propósito"],
        MODULOS, widths=[28, 30, 34, 24, 60],
        note="Los IDs de permiso se asignan por usuario en Administración → Gestión de Usuarios; el menú lateral se filtra según ellos.",
    )
    add_sheet(
        wb, "Portales", "Seven Arena — Portales de usuario final",
        ["Portal", "URL", "Quién lo usa", "Autenticación", "Funciones principales"],
        PORTALES, widths=[26, 24, 40, 30, 80],
        note="Reglas transversales: sesión única por dispositivo · eliminar mi cuenta (reversible solo desde administración) · push + campana · asistencia · ES/EN/PT.",
    )
    add_sheet(
        wb, "Roles", "Seven Arena — Roles de administración (plantillas de permisos)",
        ["Rol", "Alcance de módulos"], ROLES, widths=[22, 70],
        note="Al asignar un rol se marcan sus módulos y luego pueden ajustarse individualmente por usuario.",
    )
    out = HERE / "MODULOS_Y_PORTALES.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


def build_usuarios():
    src = HERE / "_users_export.json"
    if not src.exists():
        print("SKIP usuarios: falta _users_export.json (ejecutar: node scripts/export_users_json.js)")
        return
    data = json.loads(src.read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)

    resumen = [
        ("Usuarios de administración", len(data.get("admins", []))),
        ("Participantes (atletas/VIP/jefes)", len(data.get("athletes", []))),
        ("Conductores propios", len(data.get("drivers", []))),
        ("Participantes de proveedor", len(data.get("providerParticipants", []))),
        ("Staff de control de acceso", sum(1 for p in data.get("providerParticipants", []) if (p.get("providerType") or "").lower() == "staff")),
        ("Partners de beneficios", len(data.get("couponPartners", []))),
    ]
    add_sheet(
        wb, "Resumen", f"Seven Arena — Usuarios y códigos (corte {date.today().strftime('%d-%m-%Y')})",
        ["Tipo de usuario", "Cantidad"], resumen, widths=[42, 14],
        note="DOCUMENTO CONFIDENCIAL: contiene códigos de acceso reales. Distribuir solo al equipo autorizado.",
    )

    add_sheet(
        wb, "Administración", "Usuarios de administración (web admin)",
        ["Nombre", "Acceso (correo o usuario)", "Rol", "Módulos restringidos", "Estado", "Último ingreso"],
        [(
            a.get("name") or "—",
            (f"usuario: {a['username']}" if a.get("username") else a.get("email")),
            a.get("role") or "— (sin rol)",
            ("Sí ({} módulos)".format(len(a["modules"])) if a.get("modules") else "No (menú completo)"),
            ("Deshabilitado" if a.get("banned") else "Activo"),
            (a.get("lastSignIn") or "Nunca")[:10],
        ) for a in data.get("admins", [])],
        widths=[24, 34, 18, 24, 14, 14],
        note="Las contraseñas están cifradas y no son consultables; se restablecen desde Administración → Gestión de Usuarios (contraseña temporal con cambio obligatorio).",
    )

    add_sheet(
        wb, "Roles y Módulos", "Roles de administración (plantillas de permisos)",
        ["Rol", "Alcance de módulos"], ROLES, widths=[22, 70],
    )

    add_sheet(
        wb, "Participantes", "Participantes — portal de usuario / VIP (código = últimos 6 del ID)",
        ["Nombre", "Tipo de cliente", "Jefe de delegación", "País", "Correo", "Código de acceso", "Estado", "Portal que abre"],
        [(
            a["name"], a.get("type") or "—", "Sí" if a.get("lead") else "No",
            a.get("country") or "—", a.get("email") or "—", a["code"], a.get("status") or "—",
            ("Solicitud de Vehículo (VIP)" if a.get("type") == "VIP" else "Portal de usuario" + (" (vista jefe)" if a.get("lead") else "")),
        ) for a in data.get("athletes", [])],
        widths=[28, 14, 16, 8, 30, 16, 24, 28],
        highlight=lambda row: row[6] == "DELETED",
    )

    conductores = [(
        d["name"], "Conductor propio", "—", d.get("email") or "—", d["code"], d.get("status") or "—",
    ) for d in data.get("drivers", [])] + [(
        p["name"], "Participante de proveedor",
        f"{p.get('provider') or '—'} ({p.get('providerType') or '—'})",
        p.get("email") or "—", p["code"], p.get("status") or "—",
    ) for p in data.get("providerParticipants", []) if p.get("isDriver")]
    add_sheet(
        wb, "Conductores", "Conductores — portal del conductor (código = últimos 6 del ID)",
        ["Nombre", "Origen", "Proveedor", "Correo", "Código de acceso", "Estado"],
        conductores, widths=[26, 24, 26, 32, 16, 18],
        highlight=lambda row: row[5] == "DELETED",
    )

    staff = [(
        p["name"], f"{p.get('provider') or '—'}", p.get("email") or "—", p["code"], p.get("status") or "—",
    ) for p in data.get("providerParticipants", []) if (p.get("providerType") or "").lower() == "staff"]
    add_sheet(
        wb, "Staff Acceso", "Staff de control de acceso (proveedores tipo Staff)",
        ["Nombre", "Proveedor", "Correo", "Código de acceso", "Estado"],
        staff or [("(sin registros — crear un proveedor tipo Staff y agregarle participantes)", "—", "—", "—", "—")],
        widths=[40, 22, 30, 16, 18],
    )

    add_sheet(
        wb, "Partners", "Partners de beneficios — portal partner (código + PIN)",
        ["Comercio", "Código de login", "PIN", "Estado"],
        [(
            p["name"], p["loginCode"],
            "Cifrado — no consultable (el admin puede definir uno nuevo)",
            "Activo" if p.get("active") else "Inactivo",
        ) for p in data.get("couponPartners", [])],
        widths=[34, 18, 48, 12],
    )

    out = HERE / "USUARIOS_ROLES_Y_CODIGOS.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    build_modulos()
    build_usuarios()
